"""Phase1.7 V3 Final Review Workbook generator invariants."""
from __future__ import annotations

import csv
import importlib.util
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

BACKEND = Path(__file__).resolve().parents[1]
MOD = BACKEND / "scripts" / "phase17_v3_workbook.py"
D16 = BACKEND / "data" / "integration" / "brainregion_direct_g1_phase16"
spec = importlib.util.spec_from_file_location("phase17_v3_workbook", MOD)
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

V = "VERIFIED_DIRECT_CONTAINED"
FROZEN = "FROZEN_DECISION_PREVAILS"
ETYPE = "ONTOLOGY_ENTITY_TYPE_REVIEW"

IF_MF = {f"NGIQ-BR-{x:08d}" for x in range(371, 381)}
ZI_RT = {"NGIQ-BR-00000707", "NGIQ-BR-00000708", "NGIQ-BR-00000709", "NGIQ-BR-00000710"}
VTM_R = "NGIQ-BR-00000369"
POS = {"NGIQ-BR-00000683", "NGIQ-BR-00000684", "NGIQ-BR-00000679", "NGIQ-BR-00000680",
       "NGIQ-BR-00000677", "NGIQ-BR-00000678", "NGIQ-BR-00000717", "NGIQ-BR-00000747"}


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    out = tmp_path_factory.mktemp("wb") / "review.xlsx"
    mod.build(out)
    return load_workbook(out)


def sheet_rows(ws, start=2):
    out = []
    for r in ws.iter_rows(min_row=start, values_only=True):
        if r[0] is None:
            continue
        out.append(r)
    return out


def test_five_sheets(wb):
    names = wb.sheetnames
    assert {"01_高置信关系", "02_待复核关系", "03_关键案例_QA", "04_决策审计",
            "05_审核摘要"} <= set(names)


def test_counts(wb):
    s1 = sheet_rows(wb["01_高置信关系"])
    s2 = sheet_rows(wb["02_待复核关系"])
    assert len(s1) == 86
    assert len(s2) == 132
    assert len(sheet_rows(wb["04_决策审计"])) == 218


def test_sequential_numbers(wb):
    assert [r[0] for r in sheet_rows(wb["01_高置信关系"])] == list(range(1, 87))
    assert [r[0] for r in sheet_rows(wb["02_待复核关系"])] == list(range(1, 133))


def test_universe_membership(wb):
    all_ids = set()
    dup = []
    for ws in ("01_高置信关系", "02_待复核关系"):
        for r in sheet_rows(wb[ws]):
            eid = r[1]
            if eid in all_ids:
                dup.append(eid)
            all_ids.add(eid)
    csv_ids = {r["source_entity_id"] for r in
               csv.DictReader(open(D16 / "phase17_v3_classification.csv", encoding="utf-8-sig"))}
    assert len(all_ids) == 218
    assert all_ids == csv_ids
    assert len(dup) == 0


def test_no_fake_membership_label(wb):
    for ws in ("01_高置信关系", "02_待复核关系"):
        for r in sheet_rows(wb[ws]):
            joined = " ".join(str(x) for x in r)
            assert "NOT_IN_218_(already has path or review)" not in joined


def test_if_mf_g3_frozen_zi_rt_excluded(wb):
    s1_ids = {r[1] for r in sheet_rows(wb["01_高置信关系"])}
    assert not (s1_ids & IF_MF)
    assert not (s1_ids & ZI_RT)
    assert all(r[11] == V for r in sheet_rows(wb["01_高置信关系"]))
    # all 14 G3 frozen live only in review sheet
    s2 = sheet_rows(wb["02_待复核关系"])
    assert sum(1 for r in s2 if r[7] == FROZEN) == 14
    assert sum(1 for r in s2 if r[7] == ETYPE) == 10


def test_positive_verified_present(wb):
    s1_ids = {r[1] for r in sheet_rows(wb["01_高置信关系"])}
    assert POS <= s1_ids


def test_vtm_right_only_in_qa(wb):
    s1 = {r[1] for r in sheet_rows(wb["01_高置信关系"])}
    s2 = {r[1] for r in sheet_rows(wb["02_待复核关系"])}
    qa = {r[0] for r in sheet_rows(wb["03_关键案例_QA"])}
    assert VTM_R not in s1 and VTM_R not in s2
    assert VTM_R in qa


def test_contralateral_not_copied(wb):
    s1 = sheet_rows(wb["01_高置信关系"])
    # col21 = 对侧同源区ID (index 20); col2 = source_id (index 1)
    for r in s1:
        assert r[20] != r[1]
        assert r[20] in ("NO_PAIR",) or r[20].startswith("NGIQ-BR-")


def test_historical_conflict_loaded(wb):
    s1 = sheet_rows(wb["01_高置信关系"])
    hs = Counter(r[13] for r in s1)
    lv = Counter(r[15] for r in s1)
    assert hs.get("HISTORICAL_CONFLICT", 0) == 45
    assert hs.get("SPATIAL_COMPATIBLE", 0) == 41
    assert lv.get("G4_TO_G3", 0) == 41
    assert lv.get("SOURCE_LEVEL_SPATIAL", 0) == 45
    affects = sum(1 for r in s1 if str(r[18]) == "True" or str(r[18]).upper() == "TRUE")
    assert affects == 0


def test_review_breakdown(wb):
    s2 = sheet_rows(wb["02_待复核关系"])
    c = Counter(r[7] for r in s2)
    assert c == {FROZEN: 14, "LIKELY_CONTAINED_NEEDS_SPATIAL_REVIEW": 93,
                 ETYPE: 10, "ONTOLOGY_DEFINITION_DEPENDENT": 11,
                 "ANATOMICAL_CONFLICT": 4}


def test_qa_present_and_size(wb):
    qa = sheet_rows(wb["03_关键案例_QA"])
    assert len(qa) >= 30
    labels = {r[0] for r in qa}
    assert IF_MF <= labels
    assert {"NGIQ-BR-00000085", "NGIQ-BR-00000086"} <= labels  # STG_6_2
    assert ZI_RT <= labels


def test_no_science_change(wb):
    # generator reads frozen classification file only - no write side effects
    ids1 = {r[1] for r in sheet_rows(wb["01_高置信关系"])}
    assert len(ids1) == 86
    assert all(r[11] == V for r in sheet_rows(wb["01_高置信关系"]))


if __name__ == "__main__":
    pytest.main([__file__, "-q"])
