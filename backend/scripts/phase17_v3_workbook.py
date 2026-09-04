"""Phase1.7 V3 Final Review Workbook generator (READ-ONLY w.r.t. science).

Builds Phase1.7_V3_Final_Review.xlsx (5 sheets) from the FROZEN scientific
outputs. It never recomputes classifications - it only reads them.

Fixes the three historical V2 workbook engineering bugs at the source:
  Bug1 序号 always 1        -> sequence = enumerate(row block) starting at 1.
  Bug2 对侧 copied current  -> real contralateral entity lookup + its own fields.
  Bug3 218 membership fake  -> membership derived by set math; no
        'NOT_IN_218_(already has path or review)' fabrication.

No DB writes. No classification changes. Not committed here.
"""
from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BACKEND = Path(__file__).resolve().parents[1]
D16 = BACKEND / "data" / "integration" / "brainregion_direct_g1_phase16"
CLS = D16 / "phase17_v3_classification.csv"
HIST = D16 / "phase17_v3_historical_conflict_audit.csv"
PROV = D16 / "phase17_v3_frozen_provenance_audit.csv"
OUT = D16 / "Phase1.7_V3_Final_Review.xlsx"
PROD = "neurographiq_human_brain_v1"

V = "VERIFIED_DIRECT_CONTAINED"
FROZEN = "FROZEN_DECISION_PREVAILS"
ETYPE = "ONTOLOGY_ENTITY_TYPE_REVIEW"
ODD = "ONTOLOGY_DEFINITION_DEPENDENT"
CONFLICT = "ANATOMICAL_CONFLICT"
LIKELY = "LIKELY_CONTAINED_NEEDS_SPATIAL_REVIEW"

BUCKET_CN = {V: "高置信（VERIFIED）", FROZEN: "冻结决策优先", ETYPE: "实体类型待定",
             ODD: "本体定义依赖", CONFLICT: "解剖冲突", LIKELY: "需空间复核"}


def _csv(p):
    return list(csv.DictReader(open(p, encoding="utf-8-sig")))


def flip_name(n, gran=None):
    if n.startswith("Left "):
        return "Right " + n[5:]
    if n.startswith("Right "):
        return "Left " + n[6:]
    if n.endswith(" left"):
        return n[:-5] + " right"
    if n.endswith(" right"):
        return n[:-6] + " left"
    return None


def entity_index():
    idx = {}
    c = psycopg.connect(host="127.0.0.1", port=5432, user="postgres",
                        password="postgres", dbname=PROD, autocommit=True)
    cur = c.cursor()
    cur.execute("""SELECT b.granularity_level,e.entity_id,e.name_en
                   FROM kg_entities e JOIN brain_regions b ON b.entity_pk=e.entity_pk
                   WHERE b.granularity_level IN ('G3_MESO_FINE','G4_MICROSTRUCTURAL_FINE')""")
    for gran, eid, nm in cur.fetchall():
        idx[nm] = (eid, gran)
    c.close()
    return idx


def build(out_path: Path | None = None) -> dict:
    cls = _csv(CLS)
    hist = _csv(HIST)
    prov = _csv(PROV)
    by_id = {r["source_entity_id"]: r for r in cls}
    hist_by = {r["source_id"]: r for r in hist}
    prov_by = {r["source_entity_id"]: r for r in prov}
    ents = entity_index()
    cls_cnt = Counter(r["v3_classification"] for r in cls)
    assert len(cls) == 218 and cls_cnt[V] == 86
    if out_path is None:
        out_path = OUT

    # ---------------- contralateral pairing (real lookup) ----------------
    pair = {}
    for r in cls:
        eid = r["source_entity_id"]
        fname = flip_name(r["source_name_en"], r["source_granularity"])
        sib = ents.get(fname)
        if not sib:
            pair[eid] = dict(id="NO_PAIR", name="", cls="", g1="",
                             consistent="NO_PAIR", reason="未找到可靠对侧 canonical entity")
            continue
        sid, _ = sib
        if sid in by_id:
            s = by_id[sid]
            same_cls = (s["v3_classification"] == r["v3_classification"])
            # targets are hemisphere-specific: consistency requires the sibling's
            # target to be the mirror of this side's target (or identical)
            t1 = r["candidate_g1_name_en"]
            t2 = s["candidate_g1_name_en"]
            mirror = (flip_name(t2, "G1_PLACEHOLDER_ignore") == t1) or (t1 == t2) or \
                     (flip_name(t1, "G1_PLACEHOLDER_ignore") == t2)
            consistent = same_cls and (mirror or r["candidate_g1_entity_id"] == s["candidate_g1_entity_id"])
            pair[eid] = dict(id=sid, name=s["source_name_en"],
                             cls=s["v3_classification"], g1=s["candidate_g1_name_en"],
                             consistent="TRUE" if consistent else "FALSE",
                             reason=("两侧分类一致且 G1 目标为镜像/相同" if consistent else
                                     f"对侧 classification={s['v3_classification']} / "
                                     f"target={s['candidate_g1_name_en']}，与本侧不同"))
        else:
            pair[eid] = dict(id=sid, name=fname, cls="SIBLING_NOT_IN_218",
                             g1="", consistent="FALSE",
                             reason="对侧 canonical 存在但不在 218 universe（其状态见各自审核记录）")

    # ---------------- 01 verified ----------------
    verified = [r for r in cls if r["v3_classification"] == V]
    rows1 = []
    for i, r in enumerate(verified, start=1):
        h = hist_by.get(r["source_entity_id"], {})
        p = pair[r["source_entity_id"]]
        fam = ("丘脑" if "Thalamus" in r["candidate_g1_name_en"] else
               ("海马" if "Hippocampus" in r["candidate_g1_name_en"] else
                ("杏仁核" if "Amygdala" in r["candidate_g1_name_en"] else "其他")))
        rows1.append([
            i, r["source_entity_id"], r["source_name_zh"] or r["source_name_en"],
            r["source_name_en"],
            "右" if r["hemisphere"] == "right" else ("左" if r["hemisphere"] == "left" else r["hemisphere"]),
            "G4" if r["source_granularity"] == "G4_MICROSTRUCTURAL_FINE" else "G3",
            "Julich-Brain" if r["source_granularity"] == "G4_MICROSTRUCTURAL_FINE" else "Brainnetome",
            r["candidate_g1_entity_id"], r["candidate_g1_name_zh"] or r["candidate_g1_name_en"],
            r["candidate_g1_name_en"], "解剖包含（contained_in 候选）", "VERIFIED_DIRECT_CONTAINED", "高",
            h.get("historical_status", ""), h.get("historical_conflict_exists", ""),
            h.get("conflict_level", ""), h.get("historical_conflict_target_name", "") or "(无 G1 直接对象)",
            h.get("historical_conflict_relation_type", "") + "/" + h.get("historical_conflict_decision", ""),
            h.get("conflict_affects_current_g1_relation", ""), h.get("resolution_reason", ""),
            p["id"], p["name"], p["cls"], p["g1"], p["consistent"], p["reason"],
            r.get("anatomical_reason", ""),
            "主要证据：Julich-Brain/FreeSurfer aseg；历史证据见决策审计",
            ("家族：" + fam) if fam != "其他" else ""])

    # ---------------- 02 review ----------------
    review = [r for r in cls if r["v3_classification"] != V]
    why = {FROZEN: "存在直接针对该 G3→G1 的 frozen 决策（dominant/no-rollup/conflict），禁止生成 contained_in",
           LIKELY: "cyto↔macro 边界未被空间复核确认，不能判定为已证实的整区包含",
           ETYPE: "实体类型（FiberBundle vs BrainRegion）尚未冻结，需先确定 canonical entity type",
           ODD: "是否含于 G1 取决于 G1 ontology 定义（如 hippocampus proper 与 broader formation 的边界）",
           CONFLICT: "现代解剖定义下不能写 contained_in（如 ZI/Rt 非背侧 Thalamus）"}
    rows2 = []
    for i, r in enumerate(review, start=1):
        v3 = r["v3_classification"]
        pr = prov_by.get(r["source_entity_id"], {})
        p = pair[r["source_entity_id"]]
        rows2.append([
            i, r["source_entity_id"], r["source_name_zh"] or r["source_name_en"], r["source_name_en"],
            "右" if r["hemisphere"] == "right" else "左",
            "G4" if r["source_granularity"] == "G4_MICROSTRUCTURAL_FINE" else "G3",
            r["candidate_g1_name_en"] or "", v3, BUCKET_CN.get(v3, v3), why.get(v3, ""),
            r.get("frozen_decision", ""), pr.get("frozen_target_name", "") or r.get("frozen_target_id", ""),
            r.get("frozen_applicability", ""), r.get("frozen_rollup_eligible", ""),
            "SPATIAL_REVIEW_REQUIRED" if v3 == LIKELY else ("FROZEN" if v3 == FROZEN else "n/a"),
            "本体定义依赖" if v3 == ODD else "", "实体类型待定" if v3 == ETYPE else "",
            "完成空间复核后再判定" if v3 == LIKELY else ("重新人工审核后处理" if v3 in (FROZEN, CONFLICT) else
            "更新 ontology/entity-type 定义"), p["id"] + "/" + p["cls"] if p["id"] != "NO_PAIR" else "NO_PAIR",
            p["reason"], ""])

    # ---------------- 03 QA ----------------
    qa_ids = ["NGIQ-BR-00000005", "NGIQ-BR-00000006", "NGIQ-BR-00000085", "NGIQ-BR-00000086",
              "NGIQ-BR-00000211", "NGIQ-BR-00000212", "NGIQ-BR-00000213", "NGIQ-BR-00000214",
              "NGIQ-BR-00000125", "NGIQ-BR-00000126", "NGIQ-BR-00000131", "NGIQ-BR-00000132",
              "NGIQ-BR-00000145", "NGIQ-BR-00000146"] + \
             [f"NGIQ-BR-{x:08d}" for x in range(371, 381)] + \
             ["NGIQ-BR-00000370", "NGIQ-BR-00000369", "NGIQ-BR-00000709", "NGIQ-BR-00000710",
              "NGIQ-BR-00000707", "NGIQ-BR-00000708",
              "NGIQ-BR-00000361", "NGIQ-BR-00000364", "NGIQ-BR-00000362", "NGIQ-BR-00000365",
              "NGIQ-BR-00000363", "NGIQ-BR-00000366", "NGIQ-BR-00000367", "NGIQ-BR-00000368",
              "NGIQ-BR-00000683", "NGIQ-BR-00000677", "NGIQ-BR-00000717", "NGIQ-BR-00000747",
              "NGIQ-BR-00000725", "NGIQ-BR-00000723", "NGIQ-BR-00000735"]
    rows3 = []
    for eid in qa_ids:
        if eid == "NGIQ-BR-00000369":
            in218, cls_now, allow = "FALSE", "SEMANTIC_HIGH_SPATIAL_REVIEW(Phase1.6)", "候选（不自动 VERIFIED）"
            note = ("VTM right 原属 Phase1.6 SEMANTIC_HIGH_SPATIAL_REVIEW，不是 Phase1.6 DIRECT 218 候选；"
                    "仅进入 QA，不计入 218 统计。")
        else:
            r = by_id.get(eid, {})
            if not r:
                continue
            in218 = "TRUE"
            cls_now = r["v3_classification"]
            allow = ("是（作为 candidate 保留）" if cls_now in (V, LIKELY) else "否")
            note = ""
        rows3.append([eid, by_id.get(eid, {}).get("source_name_en", "") or eid,
                      "为什么重要/历史问题", "当前科学判断见其分类", in218, cls_now, allow,
                      "否" if cls_now in (FROZEN, CONFLICT) else "待后续", "主要证据：frozen 决策 + 对象级审计",
                      note])
    # fill QA 'why' with static text list placeholder enriched by category
    # ---------------- 04 decision audit (218) ----------------
    rows4 = []
    for r in sorted(cls, key=lambda x: x["source_entity_id"]):
        rows4.append([r["source_entity_id"], r["candidate_g1_entity_id"],
                      r["phase17_classification"], r["v3_classification"],
                      "ENTITY_TYPE_GATE" if r["gate"] == "ENTITY_TYPE_GATE" else "NONE",
                      "FROZEN_DECISION_GATE" if r["gate"] == "FROZEN_DECISION_GATE" else "NONE",
                      r.get("frozen_applicability", ""),
                      hist_by.get(r["source_entity_id"], {}).get("conflict_level", ""),
                      hist_by.get(r["source_entity_id"], {}).get("conflict_affects_current_g1_relation", ""),
                      r["gate_reason"] or r.get("phase17_classification", ""),
                      BUCKET_CN.get(r["v3_classification"], r["v3_classification"])])
    # ---------------- write workbook ----------------
    wb = Workbook()
    thin = Side(style="thin", color="C6D2DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    hdr = PatternFill("solid", fgColor="DCE6F1")
    fills = {"VERIFIED_DIRECT_CONTAINED": "E2EFDA", FROZEN: "FCE4D6",
             LIKELY: "FFF2CC", ETYPE: "F2DCDB", ODD: "E4DFEC", CONFLICT: "F8CBAD"}
    widths = dict(zip(range(1, 30), [7, 22, 30, 40, 6, 6, 14, 22, 26, 38, 22, 20, 8, 18, 12, 12,
                                     14, 40, 16, 44, 22, 30, 16, 26, 14, 44, 40, 20, 14]))

    def sheet(ws, headers, rows, status_col=None):
        ws.append(headers)
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = Font(bold=True); cell.fill = hdr; cell.alignment = cen; cell.border = border
        for row in rows:
            ws.append(row)
            rr = ws.max_row
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=rr, column=ci)
                cell.alignment = cen if ci in (1, 2, 5, 6, 7, 13, 15, 16, 17) else left
                cell.border = border
            if status_col:
                val = row[status_col - 1]
                f = fills.get(str(val))
                if f:
                    ws.cell(row=rr, column=status_col).fill = PatternFill("solid", fgColor=f)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        for ci, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(ci, 22)

    h1 = ["序号", "source_id", "源脑区中文名", "源脑区名称", "左右侧", "源粒度", "源Atlas",
          "G1目标ID", "G1目标中文名", "G1目标名称", "当前关系", "当前分类", "置信等级",
          "historical_status", "历史冲突是否存在", "历史冲突层级", "历史冲突对象", "历史冲突关系/决策",
          "是否影响当前G1关系", "历史冲突处理理由", "对侧同源区ID", "对侧同源区名称", "对侧当前分类",
          "对侧G1目标", "对侧一致性", "对侧差异说明", "科学依据摘要", "主要证据来源", "审核备注"]
    h2 = ["序号", "source_id", "源脑区中文名", "源脑区名称", "左右侧", "源粒度", "候选G1目标", "当前分类",
          "Review大类", "为什么未进入VERIFIED", "frozen decision", "frozen target", "frozen scope",
          "rollup_eligible", "空间审核状态", "本体问题", "实体类型问题", "建议下一步", "对侧状态", "备注"]
    h3 = ["source_id", "源脑区名称", "为什么重要", "当前科学判断", "是否进入218 universe",
          "当前classification", "是否允许contained_in", "是否允许rollup", "主要证据/依据", "需要导师重点看的问题"]
    h4 = ["source_id", "candidate_g1", "original_phase17_v2_classification", "v3_classification",
          "entity_type_gate", "frozen_gate", "frozen_applicability", "historical_conflict_level",
          "historical_conflict_affects_g1", "final_reason", "final_bucket"]
    ws1 = wb.active; ws1.title = "01_高置信关系"
    sheet(ws1, h1, rows1, status_col=12)
    ws2 = wb.create_sheet("02_待复核关系")
    sheet(ws2, h2, rows2, status_col=8)
    ws3 = wb.create_sheet("03_关键案例_QA")
    sheet(ws3, h3, rows3)
    ws4 = wb.create_sheet("04_决策审计")
    sheet(ws4, h4, rows4, status_col=4)
    ws5 = wb.create_sheet("05_审核摘要")
    lines = [
        ("Phase1.7 V3 审核摘要", True),
        (f"候选总数：218　　高置信(VERIFIED)：86　　待复核：132", False),
        ("待复核：FROZEN_DECISION_PREVAILS=14 · LIKELY_CONTAINED_NEEDS_SPATIAL_REVIEW=93 · "
         "ONTOLOGY_ENTITY_TYPE_REVIEW=10 · ONTOLOGY_DEFINITION_DEPENDENT=11 · ANATOMICAL_CONFLICT=4", False),
        ("- 86 条均通过 relation-level Frozen Gate（无 direct G4→G1 frozen 冲突）。", False),
        ("- historical conflict 对象级审计：G4_TO_G3=41 · SOURCE_LEVEL_SPATIAL=45 · direct G4→G1=0。", False),
        ("- IF/MF 10 条因 entity type 未冻结不进入 VERIFIED。", False),
        ("- G3 14 条继续尊重 frozen decision。", False),
        ("- ZI/Rt 不作为 Thalamus contained_in。", False),
        ("- hippocampal formation 边界项继续 ontology review。", False),
        ("- 本文件为科学审核文件，不代表已写入 canonical KG；数据库未修改。", False),
    ]
    for i, (t, bold) in enumerate(lines, start=1):
        c = ws5.cell(row=i, column=1, value=t)
        c.font = Font(bold=bold, size=13 if bold else 11)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws5.column_dimensions["A"].width = 140
    wb.save(out_path)
    return dict(verified=len(rows1), review=len(rows2), qa=len(rows3),
                decision=len(rows4), cls=dict(cls_cnt))


if __name__ == "__main__":
    res = build()
    print("saved", OUT)
    print(res)
